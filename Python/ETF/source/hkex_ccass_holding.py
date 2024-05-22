from datetime import datetime, timezone, timedelta
import time
import requests
from bs4 import BeautifulSoup
import time
import openpyxl
import pymongo
import os
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

mongodb_url=os.environ.get('CONFIG_MONGODB_URL', 'mongodb://testing:testing@192.168.0.9:27017/')

client=pymongo.MongoClient(mongodb_url)
db=client["etf_db"]
table=db["etf_aum_table"]
latest_table=db["etf_latest_aum_table"]

codes=["03008","09008","03009","09009","03042","83042","09042","03046","83046","09046","03439","09439","03179","09179","03066","03068","03135"]

def extract_content(stock_code,to_date):
    site="https://www3.hkexnews.hk/"
    URL=f"{site}/sdw/search/searchsdw.aspx?__EVENTTARGET=btnSearch&txtShareholdingDate={to_date}&txtStockCode={stock_code}"
    print(URL)
    
    page=requests.get(URL)
    soup=BeautifulSoup(page.content,"html.parser")
    
    search_date=soup.find(id="txtShareholdingDate")
    print(search_date)
    print(search_date["value"])
    
    search_stock=soup.find(id="txtStockName")
    print(search_stock)
    print(search_stocck["value")
    
    
    summary_value=soup.find("div",class_="summary-value")
    print(summary_value)
    print(summary_value.text)
    
    ind_value=soup.find("div",class_="ccass-search-total").find("div",class_="value")
    print(ind_value)
    print(ind_value.text)
    


def parse_file(stock_code,link):
    file=f"{stock_code}.xlsx"
    wb=openpyxl.load_workbook(file)
    sheet=wb.active
    for col in [3,6,9,12,15]:
        if (sheet.cell(row=8,column=col).value):
            code=sheet.cell(row=8,column=col).value
            if str(code) in codes:
                report_date=sheet.cell(row=10,column=col).value.strftime("%Y%m%d")

                total_units_outstanding=sheet.cell(row=18,column=col).value
                ccy=sheet.cell(row=21,column=(col-1)).value
                aum=sheet.cell(row=21,column=col).value
                result={
                    "Date":report_date,
                    "Stock":code,
                    "AUM":aum,
                    "Units":total_units_outstanding,
                    "Currency":ccy,
                    "Z-Link":link
                    
                }
                key={
                    "Date":report_date,
                    "Stock":code
                }
                update_latest_key={
                    "Stock":code
                }
                table.update_many(key,{"$set": result},upsert=True)
                latest_table.update_many(update_latest_key,{"$set": result},upsert=True)
                


def trigger_job():
    to_date=datetime.fromtimestamp(int(time.time()-24*3600)).strftime('%Y%m%d')
    to_date="20240521"
    stocks={}
    stocks["3008"]="1000221537"
 
    
    for stock in stocks.keys():
        extract_content(stock,to_date)
        time.sleep(5)
    

if __name__ == "__main__":
    sched = BackgroundScheduler(
        timezone='UTC',
        job_defaults={'misfire_grace_time': 15*60}
        )
    
    trigger = CronTrigger(
        ### HKT 16,17,19,23,7
        year="*", month="*", day="*", hour="8,9,11,15,23", minute="45", second="0"
    )
    sched.add_job(
        trigger_job,
        trigger=trigger,
        args=[],
        name="trigger job",
    )  
    #sched.add_job(job2,'interval',id='2_sec',seconds=15)
    
    trigger_job()
    