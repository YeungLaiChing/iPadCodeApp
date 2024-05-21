from datetime import datetime, timezone, timedelta
import time
import requests
from bs4 import BeautifulSoup
import time
import openpyxl
import pymongo
import os
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

mongodb_url=os.environ.get('CONFIG_MONGODB_URL', 'mongodb://testing:testing@192.168.0.9:27017/')

client=pymongo.MongoClient(mongodb_url)
db=client["etf_db"]
table=db["etf_aum_table"]
latest_table=db["etf_latest_aum_table"]

codes=["3008","9008","3009","9009","3042","83042","9042","3046","83046","9046","3439","9439","3179","9179","3066","3068","3135"]

def download_first_file(stock_code,stock_id,from_date,to_date):
    site="https://www1.hkexnews.hk/"
    URL=f"{site}/search/titlesearch.xhtml?lang=EN&category=0&market=SEHK&searchType=1&documentType=-1&t1code=80000&t2Gcode=-2&t2code=-2&stockId={stock_id}&from={from_date}&to={to_date}&MB-Daterange=0&title="
    page=requests.get(URL)
    soup=BeautifulSoup(page.content,"html.parser")
    elements=soup.find_all("div",class_="doc-link")
    link=""

    for element in elements:
        content=element.find("a")
        if content["href"]>link:
            link=content["href"]
        
    download_url=f"{site}/{link}"

    print(download_url)

    filename=f"{stock_code}.xlsx"

    resp=requests.get(download_url)
    if resp.ok:
        with open(filename,mode="wb") as file:
            file.write(resp.content)
    return download_url

def parse_file(stock_code,link):
    file=f"{stock_code}.xlsx"
    wb=openpyxl.load_workbook(file)
    sheet=wb.active
    for col in [3,6,9,12,15]:
        if (sheet.cell(row=8,column=col).value):
            code=sheet.cell(row=8,column=col).value
            if str(code) in codes:
                report_date=sheet.cell(row=10,column=col).value.strftime("%Y%m%d")

                total_units_outstanding=sheet.cell(row=18,column=col).value
                ccy=sheet.cell(row=21,column=(col-1)).value
                aum=sheet.cell(row=21,column=col).value
                result={
                    "Date":report_date,
                    "Stock":code,
                    "AUM":aum,
                    "Units":total_units_outstanding,
                    "Currency":ccy,
                    "Z-Link":link
                    
                }
                key={
                    "Date":report_date,
                    "Stock":code
                }
                update_latest_key={
                    "Stock":code
                }
                table.update_many(key,{"$set": result},upsert=True)
                latest_table.update_many(update_latest_key,{"$set": result},upsert=True)
                


def trigger_job():
    to_date=datetime.fromtimestamp(int(time.time())).strftime('%Y%m%d')
    from_date=datetime.fromtimestamp(int(time.time())-10*3600*24).strftime('%Y%m%d')
    stocks={}
    stocks["3008"]="1000221537"
    stocks["3042"]="1000221230"
    stocks["3439"]="1000221233"
    stocks["3179"]="1000221232"
    stocks["3046"]="1000221231"
    stocks["3009"]="1000221538"
    
    stocks["3066"]="1000179272"
    
    stocks["3135"]="1000181396"
    
    for stock in stocks.keys():
        lnk=download_first_file(stock,stocks[stock],from_date,to_date)
        parse_file(stock,lnk)
        time.sleep(5)
    

if __name__ == "__main__":
    sched = BackgroundScheduler(
        timezone='Asia/Hong_Kong',
        job_defaults={'misfire_grace_time': 15*60}
        )
    
    trigger = CronTrigger(
        year="*", month="*", day="*", hour="8,9,10,12,13,14,17,18,19,23", minute="12", second="0"
    )
    sched.add_job(
        trigger_job,
        trigger=trigger,
        args=[],
        name="trigger job",
    )  
    #sched.add_job(job2,'interval',id='2_sec',seconds=15)
    time.sleep(5)
    sched.start()
    while True:
        time.sleep(10000)
    