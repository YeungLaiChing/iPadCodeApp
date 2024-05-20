import json
import requests
from bs4 import BeautifulSoup
import time
import openpyxl

codes=["3008","9008","3009","9009","3042","83042","9042","3046","83046","9046","3439","9439","3179","9179","3066","3068","3135"]

def download_first_file(stock_code,stock_id,from_date,to_date):
    site="https://www1.hkexnews.hk/"
    URL=f"{site}/search/titlesearch.xhtml?lang=EN&category=0&market=SEHK&searchType=1&documentType=-1&t1code=80000&t2Gcode=-2&t2code=-2&stockId={stock_id}&from={from_date}&to={to_date}&MB-Daterange=0&title="
    page=requests.get(URL)
    soup=BeautifulSoup(page.content,"html.parser")
    elements=soup.find_all("div",class_="doc-link")
    link=""

    for element in elements:
        content=element.find("a")
        if content["href"]>link:
            link=content["href"]
        
    download_url=f"{site}/{link}"

    print(download_url)

    filename=f"{stock_code}.xlsx"

    resp=requests.get(download_url)
    if resp.ok:
        with open(filename,mode="wb") as file:
            file.write(resp.content)

def parse_file(stock_code):
    file=f"{stock_code}.xlsx"
    wb=openpyxl.load_workbook(file)
    sheet=wb.active
    for col in [3,6,9,12,15]:
        if (sheet.cell(row=8,column=col).value):
            code=sheet.cell(row=8,column=col).value
            if str(code) in codes:
                report_date=sheet.cell(row=10,column=col).value
                total_units_outstanding=sheet.cell(row=18,column=col).value
                ccy=sheet.cell(row=21,column=(col-1)).value
                aum=sheet.cell(row=21,column=col).value
                print(f"{code},{report_date},{total_units_outstanding},{ccy},{aum}")

def run():
    
    stocks={}
    stocks["3008"]="1000221537"
    stocks["3042"]="1000221230"
    stocks["3439"]="1000221233"
    stocks["3179"]="1000221232"
    stocks["3046"]="1000221231"
    stocks["3009"]="1000221538"
    
    stocks["3066"]="1000179272"
    
    stocks["3135"]="1000181396"
   
    for stock in stocks.keys():
        download_first_file(stock,stocks[stock],"20240517","20240520")
        parse_file(stock)
        time.sleep(5)
    

if __name__ == "__main__":
    run()